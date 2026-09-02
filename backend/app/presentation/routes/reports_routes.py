from flask import Blueprint, send_file, request, jsonify
from flask_jwt_extended import jwt_required, get_jwt_identity
from app.infrastructure.database.models.models import Project, KPIMetric, db, User, AuditLog, KnowledgeRepository, ProjectMember
from app.utils.report_gen import generate_excel_report, generate_pdf_summary
import io
import csv
import zipfile

from app.domain.services.feature_engine import feature_module_required
from app.domain.services.document_branding_service import DocumentBrandingService

reports_bp = Blueprint('reports', __name__)

def _get_user_projects(user):
    """Return the Project query filtered to what this user can see."""
    role = user.role.name if user.role else 'Team Member'
    query = Project.query.filter_by(org_id=user.org_id)
    if role in ('Admin', 'CEO', 'SuperAdmin'):
        pass
    elif role == 'Facilitator':
        query = query.filter(Project.facilitator_id == user.id)
    elif role == 'Reviewer':
        query = query.filter(Project.reviewer_id == user.id)
    elif role == 'Team Leader':
        query = query.filter(db.or_(
            Project.team_leader_id == user.id,
            Project.creator_id == user.id,
            Project.members.any(id=user.id)
        ))
    elif role == 'Team Member':
        query = query.filter(Project.members.any(id=user.id))
    else:
        query = query.filter(False)
    return query


@reports_bp.route('/export/excel', methods=['GET'])
@jwt_required()
@feature_module_required('reports.excel')
def export_excel():
    user_id = get_jwt_identity()
    user = db.session.get(User, user_id)
    if not user:
        return jsonify({"msg": "User not found"}), 404
        
    projects = _get_user_projects(user).all()
    excel_file = generate_excel_report(projects)
    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='projects_report.xlsx'
    )


@reports_bp.route('/export/csv', methods=['GET'])
@jwt_required()
@feature_module_required('reports.csv')
def export_csv():
    """Export all accessible projects as a CSV executive summary."""
    user_id = int(get_jwt_identity())
    user = db.session.get(User, user_id)
    if not user:
        return jsonify({"msg": "User not found"}), 404

    projects = (
        _get_user_projects(user)
        .options(
            db.joinedload(Project.department),
            db.joinedload(Project.team_leader),
            db.joinedload(Project.facilitator)
        )
        .all()
    )

    db.session.add(AuditLog(
        org_id=user.org_id,
        user_id=user.id,
        action="EXPORT_CSV",
        target_table="projects",
        details={"count": len(projects), "ip": request.remote_addr}
    ))
    db.session.commit()

    proj_ids = [p.id for p in projects]

    # Batch fetch Stage 7 ROI workflow records (1 query instead of N queries)
    s7_roi_by_proj = {}
    if proj_ids:
        try:
            from app.infrastructure.database.models.models import ProjectWorkflow
            wf7_list = ProjectWorkflow.query.filter(
                ProjectWorkflow.project_id.in_(proj_ids),
                ProjectWorkflow.stage_id == 7
            ).all()
            for wf7 in wf7_list:
                if wf7 and wf7.data and isinstance(wf7.data, dict):
                    s7_roi_by_proj[wf7.project_id] = wf7.data.get('roi_validation') or {}
        except Exception:
            pass

    # Batch fetch ProjectStageTracker records (1 query instead of N queries)
    trackers_by_proj = {}
    if proj_ids:
        try:
            from app.infrastructure.database.models.models import ProjectStageTracker
            all_trackers = ProjectStageTracker.query.filter(
                ProjectStageTracker.project_id.in_(proj_ids)
            ).all()
            for t in all_trackers:
                trackers_by_proj.setdefault(t.project_id, []).append(t)
        except Exception:
            pass

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        'Project UID', 'Title', 'Department', 'Status', 'Priority',
        'Start Date', 'End Date', 'Team Leader', 'Facilitator',
        'Completion %', 'Annual Savings (INR)', 'Investment (INR)',
        'ROI %', 'Payback Period'
    ])

    for p in projects:
        s7_roi = s7_roi_by_proj.get(p.id, {})
        trackers = trackers_by_proj.get(p.id, [])
        completed = sum(1 for t in trackers if str(t.status).lower() == 'completed')
        total_stages = max(len(trackers), 1)
        comp_pct = round((completed / total_stages) * 100) if trackers else 0

        dept_name = p.department.name if p.department else '--'
        tl = p.team_leader
        tl_name = (tl.full_name or f"{getattr(tl, 'first_name', '')} {getattr(tl, 'last_name', '')}".strip() or tl.username) if tl else '--'
        fac = p.facilitator
        fac_name = (fac.full_name or f"{getattr(fac, 'first_name', '')} {getattr(fac, 'last_name', '')}".strip() or fac.username) if fac else '--'

        writer.writerow([
            p.project_uid or f'PRJ-{p.id}',
            p.title or '--',
            dept_name,
            p.status or '--',
            getattr(p, 'priority', '--') or '--',
            p.start_date.strftime('%Y-%m-%d') if p.start_date else '--',
            p.end_date.strftime('%Y-%m-%d') if p.end_date else '--',
            tl_name,
            fac_name,
            f'{comp_pct}%',
            s7_roi.get('annual_savings', '--'),
            s7_roi.get('total_investment', '--'),
            s7_roi.get('roi_pct', '--'),
            s7_roi.get('payback_period', '--'),
        ])

    csv_bytes = output.getvalue().encode('utf-8-sig')
    return send_file(
        io.BytesIO(csv_bytes),
        mimetype='text/csv',
        as_attachment=True,
        download_name='QCMS_Analytics_Report.csv'
    )


@reports_bp.route('/export/pdf/all', methods=['GET'])
@jwt_required()
@feature_module_required('reports.pdf')
def export_all_pdfs():
    """Generate PDF for every accessible project and return as a ZIP archive."""
    user_id = int(get_jwt_identity())
    user = db.session.get(User, user_id)
    if not user:
        return jsonify({"msg": "User not found"}), 404

    is_super_admin = bool(user.role and user.role.name == 'SuperAdmin')
    closed_statuses = ('Closed', 'Completed', 'Archived', 'Stage 8 Approved', 'Stage 8 Submitted', 'Pending Closure', 'SOP Created')

    if is_super_admin:
        projects = [p for p in Project.query.all() if p.status in closed_statuses]
    else:
        projects = [p for p in _get_user_projects(user).all() if p.status in closed_statuses]

    if not projects:
        return jsonify({"msg": "No closed or completed projects available for bulk report download"}), 404

    from app.utils.pdf_filler import generate_qc_story_closure_summary_pdf

    zip_buffer = io.BytesIO()
    success_count = 0
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        for p in projects:
            try:
                pdf_data = generate_qc_story_closure_summary_pdf(p.id)
                if not pdf_data:
                    from app.utils.report_gen import generate_pdf_summary
                    from app.infrastructure.database.models.models import KPIMetric
                    kpi = KPIMetric.query.filter_by(project_id=p.id).first()
                    pdf_out = generate_pdf_summary(p, kpi, p.org_id)
                    if pdf_out:
                        pdf_data = pdf_out.encode('latin-1') if isinstance(pdf_out, str) else bytes(pdf_out)

                if pdf_data:
                    safe_uid = (p.project_uid or f'PRJ_{p.id}').replace('/', '_')
                    zf.writestr(f"{safe_uid}_QC_Story_Report.pdf", pdf_data)
                    success_count += 1
            except Exception as e:
                print(f"[PDF Export] Skipped project {p.id}: {e}")

    if success_count == 0:
        return jsonify({"msg": "No PDFs could be generated"}), 500

    db.session.add(AuditLog(
        org_id=user.org_id,
        user_id=user.id,
        action="EXPORT_PDF_ALL",
        target_table="projects",
        details={"count": success_count, "ip": request.remote_addr}
    ))
    db.session.commit()

    zip_buffer.seek(0)
    return send_file(
        zip_buffer,
        mimetype='application/zip',
        as_attachment=True,
        download_name='QCMS_All_Project_Reports.zip'
    )


# ==============================================================================
# [DEAD CODE - UNUSED BY FRONTEND / REMOVED FEATURE]
# Function: export_all_pdfs_async (Lines 228-299)
# Reason: Celery async bulk PDF generation. Frontend uses synchronous /export/pdf/all.
# ==============================================================================
# @reports_bp.route('/export/pdf/all-async', methods=['POST', 'GET'])
# @reports_bp.route('/projects/export-pdf-all-async', methods=['POST', 'GET'])
# @jwt_required()
# @feature_module_required('reports.pdf')
# def export_all_pdfs_async():
#     """
#     Spawns bulk project PDF generation & ZIP packaging in Celery background queue.
#     Returns HTTP 202 Accepted with a polling job_id immediately.
#     """
#     user_id = int(get_jwt_identity())
#     user = db.session.get(User, user_id)
#     if not user:
#         return jsonify({"msg": "User not found"}), 404

#     is_super_admin = bool(user.role and user.role.name == 'SuperAdmin')
#     job_id = f"job_bulk_zip_{user.org_id or 'sa'}_{int(datetime.now(timezone.utc).replace(tzinfo=None).timestamp())}_{uuid.uuid4().hex[:6]}"

#     job_data = {
#         "job_id": job_id,
#         "org_id": user.org_id,
#         "status": "processing",
#         "progress": 10,
#         "created_at": datetime.now(timezone.utc).replace(tzinfo=None).timestamp(),
#         "completed_at": None,
#         "download_url": None,
#         "filename": None,
#         "error": None
#     }
#     _set_pdf_job(job_id, job_data)

#     try:
#         from app.infrastructure.tasks.report_tasks import generate_async_bulk_pdf_zip
#         generate_async_bulk_pdf_zip.apply_async(
#             args=[user.id, user.org_id, is_super_admin],
#             task_id=job_id,
#             retry=False
#         )
#     except Exception as celery_err:
#         current_app.logger.warning(f"[Async Bulk PDF] Celery dispatch fallback: {celery_err}")
#         app_obj = current_app._get_current_object()
#         def _bg_bulk_worker(app_inst, u_id, o_id, is_sa, jid):
#             with app_inst.app_context():
#                 try:
#                     from app.infrastructure.tasks.report_tasks import generate_async_bulk_pdf_zip
#                     res = generate_async_bulk_pdf_zip.run(user_id=u_id, org_id=o_id, is_super_admin=is_sa)
#                     j = _get_pdf_job(jid) or {}
#                     if res.get("status") == "completed":
#                         j["status"] = "completed"
#                         j["progress"] = 100
#                         j["download_url"] = res.get("download_url")
#                         j["filename"] = res.get("filename")
#                         j["completed_at"] = res.get("completed_at")
#                     else:
#                         j["status"] = "failed"
#                         j["error"] = res.get("error", "Failed to generate bulk PDF zip")
#                     _set_pdf_job(jid, j)
#                 except Exception as err:
#                     j = _get_pdf_job(jid) or {}
#                     j["status"] = "failed"
#                     j["error"] = str(err)
#                     _set_pdf_job(jid, j)

#         t = threading.Thread(target=_bg_bulk_worker, args=(app_obj, user.id, user.org_id, is_super_admin, job_id))
#         t.daemon = True
#         t.start()

#     return jsonify({
#         "status": "processing",
#         "job_id": job_id,
#         "message": "Bulk project PDF generation & ZIP packaging started in background.",
#         "poll_url": f"/api/reports/jobs/{job_id}"
#     }), 202
# [END DEAD CODE: export_all_pdfs_async]



@reports_bp.route('/export/pdf/<int:project_id>', methods=['GET'])
@jwt_required()
@feature_module_required('reports.pdf')
def export_pdf(project_id):
    user_id = int(get_jwt_identity())
    user = db.session.get(User, user_id)
    if not user:
        return jsonify({"msg": "User not found"}), 404

    is_super_admin = bool(user.role and user.role.name == 'SuperAdmin')

    # Resolve project by project_id or KnowledgeRepository entry ID
    if is_super_admin:
        project = Project.query.filter_by(id=project_id).first()
        if not project:
            kr = KnowledgeRepository.query.filter_by(id=project_id).first()
            if kr:
                project = Project.query.filter_by(id=kr.project_id).first()
    else:
        project = Project.query.filter_by(id=project_id, org_id=user.org_id).first()
        if not project:
            kr = KnowledgeRepository.query.filter_by(id=project_id, org_id=user.org_id).first()
            if kr:
                project = Project.query.filter_by(id=kr.project_id, org_id=user.org_id).first()
    
    if not project:
        return jsonify({"msg": "Project not found"}), 404

    is_admin = user.role and user.role.name in ('Admin', 'CEO', 'SuperAdmin')
    is_assigned = (project.creator_id == user.id) or (project.team_leader_id == user.id) or (project.facilitator_id == user.id) or (project.reviewer_id == user.id) or (ProjectMember.query.filter_by(project_id=project.id, user_id=user.id).first() is not None)
    is_archived = KnowledgeRepository.query.filter_by(project_id=project.id, org_id=user.org_id).first() is not None
    if not (is_admin or is_assigned or is_archived):
        return jsonify({"msg": "Unauthorized to export report for this project."}), 403

    closed_statuses = ('Closed', 'Completed', 'Archived', 'Stage 8 Approved', 'Stage 8 Submitted', 'Pending Closure', 'SOP Created')
    is_closed = (project.status in closed_statuses) or is_super_admin or (user.role and user.role.name in ('Admin', 'CEO'))

    if not is_closed:
        return jsonify({"msg": "Project report generation is disabled until Stage 8 completion and project closure."}), 400

    # Enforce role-based membership for non-admin active projects
    if not is_closed and not is_super_admin:
        role = user.role.name if user.role else 'Team Member'
        if role in ('Admin', 'CEO', 'SuperAdmin'):
            pass
        elif role == 'Facilitator':
            if project.facilitator_id != user.id:
                return jsonify({"msg": "Unauthorized"}), 403
        elif role == 'Reviewer':
            if project.reviewer_id != user.id:
                return jsonify({"msg": "Unauthorized"}), 403
        elif role == 'Team Leader':
            is_member = ProjectMember.query.filter_by(project_id=project.id, user_id=user.id).first()
            if project.team_leader_id != user.id and project.creator_id != user.id and not is_member:
                return jsonify({"msg": "Unauthorized"}), 403
        elif role == 'Team Member':
            is_member = ProjectMember.query.filter_by(project_id=project.id, user_id=user.id).first()
            if not is_member:
                return jsonify({"msg": "Unauthorized"}), 403
        else:
            return jsonify({"msg": "Unauthorized"}), 403

    tool_name = request.args.get('tool')

    db.session.add(AuditLog(
        org_id=user.org_id or project.org_id,
        user_id=user.id,
        project_id=project.id,
        action=f"EXPORT_PDF_{tool_name.upper()}" if tool_name else "EXPORT_PDF_8D",
        target_table="projects",
        target_id=project.id,
        details={"project_uid": project.project_uid, "ip": request.remote_addr}
    ))
    db.session.commit()

    if tool_name:
        from app.utils.report_gen import generate_qc_tool_report
        pdf_data = generate_qc_tool_report(project.id, tool_name)
        if not pdf_data:
            return jsonify({"msg": f"Failed to generate report for tool {tool_name}"}), 400

        return send_file(
            io.BytesIO(pdf_data),
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f"{project.project_uid}_{tool_name}_report.pdf"
        )
    else:
        from app.utils.pdf_filler import generate_qc_story_closure_summary_pdf
        pdf_data = None
        try:
            pdf_data = generate_qc_story_closure_summary_pdf(project.id)
        except Exception as err:
            print(f"[PDF Export] generate_qc_story_closure_summary_pdf error: {err}")

        # Fallback to FPDF summary if needed
        if not pdf_data:
            try:
                from app.utils.report_gen import generate_pdf_summary
                from app.infrastructure.database.models.models import KPIMetric
                kpi = KPIMetric.query.filter_by(project_id=project.id).first()
                pdf_out = generate_pdf_summary(project, kpi, project.org_id)
                if pdf_out:
                    pdf_data = pdf_out.encode('latin-1') if isinstance(pdf_out, str) else bytes(pdf_out)
            except Exception as fpdf_err:
                print(f"[PDF Export] FPDF summary fallback error: {fpdf_err}")

        if not pdf_data:
            return jsonify({"msg": "Failed to generate QC Story report"}), 400

        return send_file(
            io.BytesIO(pdf_data),
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f"{project.project_uid}_QC_Story_Report.pdf"
        )


# ── Asynchronous Background PDF Report Generation ─────────────────────────────
import threading
import uuid
from datetime import datetime, timezone
from flask import current_app
from app.infrastructure.cache.redis_adapter import cache
from app.presentation.routes.error_helpers import internal_server_error

def _set_pdf_job(job_id: str, data: dict):
    cache.setex(f"pdf_job:{job_id}", 7200, data)

def _get_pdf_job(job_id: str) -> dict:
    return cache.get(f"pdf_job:{job_id}")

# ==============================================================================
# [DEAD CODE - UNUSED BY FRONTEND / REMOVED FEATURE]
# Function: export_pdf_async (Lines 430-571)
# Reason: Celery async single PDF export. Frontend uses synchronous /export/pdf/<id>.
# ==============================================================================
# @reports_bp.route('/projects/<int:project_id>/export-pdf-async', methods=['POST'])
# @reports_bp.route('/export/pdf/<int:project_id>/async', methods=['POST'])
# @jwt_required()
# @feature_module_required('reports.pdf')
# def export_pdf_async(project_id):
#     """
#     Spawns complete project PDF report generation in a background daemon thread.
#     Returns immediately with 202 Accepted and a polling job_id.
#     """
#     user_id = int(get_jwt_identity())
#     user = db.session.get(User, user_id)
#     if not user:
#         return jsonify({"msg": "User not found"}), 404

#     is_super_admin = bool(user.role and user.role.name == 'SuperAdmin')
#     if is_super_admin:
#         project = Project.query.filter_by(id=project_id).first()
#     else:
#         project = Project.query.filter_by(id=project_id, org_id=user.org_id).first()

#     if not project:
#         return jsonify({"msg": "Project not found"}), 404

#     job_id = f"job_pdf_{project_id}_{int(datetime.now(timezone.utc).replace(tzinfo=None).timestamp())}_{uuid.uuid4().hex[:6]}"
#     tool_name = request.args.get('tool') or (request.get_json(silent=True) or {}).get('tool')

#     job_data = {
#         "job_id": job_id,
#         "project_id": project_id,
#         "project_uid": project.project_uid,
#         "tool_name": tool_name,
#         "status": "processing",
#         "progress": 10,
#         "created_at": datetime.now(timezone.utc).replace(tzinfo=None).timestamp(),
#         "completed_at": None,
#         "download_url": None,
#         "filename": None,
#         "error": None
#     }
#     _set_pdf_job(job_id, job_data)

#     app = current_app._get_current_object()

#     def _async_pdf_worker(target_app, pid, jid, tool, u_id, o_id):
#         with target_app.app_context():
#             from app.infrastructure.database.models.models import Project, KPIMetric, AuditLog, db
#             from app.infrastructure.storage import storage
#             try:
#                 p = db.session.get(Project, pid)
#                 if not p:
#                     j = _get_pdf_job(jid) or {}
#                     j["status"] = "failed"
#                     j["error"] = "Project not found"
#                     _set_pdf_job(jid, j)
#                     return

#                 j = _get_pdf_job(jid) or {}
#                 j["progress"] = 40
#                 _set_pdf_job(jid, j)

#                 pdf_data = None
#                 filename = f"{p.project_uid}_QC_Story_Report.pdf"

#                 if tool:
#                     from app.utils.report_gen import generate_qc_tool_report
#                     pdf_data = generate_qc_tool_report(p.id, tool)
#                     filename = f"{p.project_uid}_{tool}_report.pdf"
#                 else:
#                     from app.utils.pdf_filler import generate_qc_story_closure_summary_pdf
#                     try:
#                         pdf_data = generate_qc_story_closure_summary_pdf(p.id)
#                     except Exception as gen_err:
#                         print(f"[Async PDF] Primary filler error: {gen_err}")

#                     if not pdf_data:
#                         from app.utils.report_gen import generate_pdf_summary
#                         kpi = KPIMetric.query.filter_by(project_id=p.id).first()
#                         pdf_out = generate_pdf_summary(p, kpi, p.org_id)
#                         if pdf_out:
#                             pdf_data = pdf_out.encode('latin-1') if isinstance(pdf_out, str) else bytes(pdf_out)

#                 if not pdf_data:
#                     j = _get_pdf_job(jid) or {}
#                     j["status"] = "failed"
#                     j["error"] = "Could not generate PDF content"
#                     _set_pdf_job(jid, j)
#                     return

#                 j = _get_pdf_job(jid) or {}
#                 j["progress"] = 80
#                 _set_pdf_job(jid, j)

#                 saved = storage.save_file(
#                     pdf_data,
#                     filename=filename,
#                     subfolder="reports",
#                     content_type="application/pdf"
#                 )

#                 j = _get_pdf_job(jid) or {}
#                 j["status"] = "completed"
#                 j["progress"] = 100
#                 j["completed_at"] = datetime.now(timezone.utc).replace(tzinfo=None).timestamp()
#                 j["download_url"] = saved.get("url")
#                 j["filename"] = saved.get("filename")
#                 _set_pdf_job(jid, j)

#                 db.session.add(AuditLog(
#                     org_id=o_id,
#                     user_id=u_id,
#                     project_id=p.id,
#                     action=f"ASYNC_EXPORT_PDF_{tool.upper()}" if tool else "ASYNC_EXPORT_PDF_8D",
#                     target_table="projects",
#                     target_id=p.id,
#                     details={"job_id": jid, "url": saved.get("url")}
#                 ))
#                 db.session.commit()
#             except Exception as e:
#                 print(f"[Async PDF Worker Error] {e}")
#     # Dispatch to Celery distributed worker queue with local fallback
#     try:
#         from app.infrastructure.tasks.report_tasks import generate_async_pdf_report
#         generate_async_pdf_report.apply_async(
#             args=[project_id, tool_name, user.id, user.org_id],
#             task_id=job_id,
#             retry=False
#         )
#     except Exception as celery_err:
#         current_app.logger.warning(f"[Async Reports] Celery dispatch fallback to direct worker: {celery_err}")
#         thread = threading.Thread(
#             target=_async_pdf_worker,
#             args=(app, project_id, job_id, tool_name, user.id, user.org_id)
#         )
#         thread.daemon = True
#         thread.start()

#     return jsonify({
#         "status": "processing",
#         "job_id": job_id,
#         "message": "Report generation started in background.",
#         "poll_url": f"/api/reports/jobs/{job_id}"
#     }), 202
# [END DEAD CODE: export_pdf_async]


# ==============================================================================
# [DEAD CODE - UNUSED BY FRONTEND / REMOVED FEATURE]
# Function: get_pdf_job_status (Lines 573-612)
# Reason: Async Celery task poller endpoint.
# ==============================================================================
# @reports_bp.route('/jobs/<string:job_id>', methods=['GET'])
# @reports_bp.route('/task-status/<string:job_id>', methods=['GET'])
# @jwt_required()
# def get_pdf_job_status(job_id):
#     """Poll the status of a background report generation job from Celery / Redis."""
#     try:
#         from celery.result import AsyncResult
#         task_res = AsyncResult(job_id)
#         if task_res.state == 'SUCCESS':
#             data = task_res.result or {}
#             return jsonify({
#                 "job_id": job_id,
#                 "status": "completed",
#                 "progress": 100,
#                 "download_url": data.get("download_url"),
#                 "filename": data.get("filename"),
#                 "completed_at": data.get("completed_at")
#             }), 200
#         elif task_res.state == 'PROGRESS':
#             meta = task_res.info or {}
#             return jsonify({
#                 "job_id": job_id,
#                 "status": "processing",
#                 "progress": meta.get("progress", 50),
#                 "message": meta.get("status", "Generating report in background...")
#             }), 200
#         elif task_res.state == 'FAILURE':
#             return jsonify({
#                 "job_id": job_id,
#                 "status": "failed",
#                 "error": "Report generation failed. Please try again."
#             }), 200
#     except Exception:
#         pass

#     # Fallback to direct Redis job cache
#     job = _get_pdf_job(job_id)
#     if not job:
#         return jsonify({"status": "not_found", "message": "Job not found or expired"}), 404
#     return jsonify(job), 200
# [END DEAD CODE: get_pdf_job_status]



@reports_bp.route('/download-mock', methods=['GET'])
@jwt_required()
def download_mock():
    from datetime import datetime
    import io
    import csv
    from app.infrastructure.database.models.models import (
        User, Organization, Subscription, SubscriptionPayment, Module, SupportTicket
    )
    
    user_id = get_jwt_identity()
    user = db.session.get(User, int(user_id)) if user_id else None
    report_type = request.args.get('type', 'dashboard')
    fmt = request.args.get('format', 'PDF').upper()
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    def clean_str(s):
        if s is None:
            return ""
        s = str(s)
        s = s.replace("—", "-")
        s = s.replace("₹", "INR ")
        s = s.replace("–", "-")
        s = s.replace("’", "'")
        s = s.replace("“", '"')
        s = s.replace("”", '"')
        return s.encode('latin-1', 'replace').decode('latin-1')

    # Gather data based on segment
    headers_list = []
    data_list = []
    summary_metrics = []
    
    if report_type == 'overview':
        total_orgs = Organization.query.filter(Organization.is_deleted == False, Organization.is_platform_org == False).count()
        active_orgs = Organization.query.filter(Organization.is_deleted == False, Organization.is_platform_org == False, Organization.subscription_status == 'Active').count()
        total_users = User.query.join(Organization, User.org_id == Organization.id).filter(Organization.is_deleted == False, Organization.is_platform_org == False).count()
        active_users = User.query.join(Organization, User.org_id == Organization.id).filter(User.is_active == True, Organization.is_deleted == False, Organization.is_platform_org == False).count()
        total_revenue = db.session.query(db.func.sum(SubscriptionPayment.final_amount)).join(Organization, SubscriptionPayment.org_id == Organization.id).filter(Organization.is_platform_org == False, SubscriptionPayment.payment_status == 'Completed').scalar() or 0.0
        
        summary_metrics = [
            ["Total Active Organizations", f"{total_orgs} Orgs", "Public / Tenant"],
            ["Paid Active Subscriptions", f"{active_orgs} Subscriptions", "Financial / Admin"],
            ["Total Registered Accounts", f"{total_users} Users", "Identity / User"],
            ["Active User Accounts", f"{active_users} Users", "Identity / User"],
            ["Platform Ledger Revenue", f"INR {total_revenue:,.2f}", "Financial / Private"]
        ]
        
        headers_list = ["Organization Name", "Industry", "Plan", "Status", "Joined On"]
        recent_orgs = Organization.query.filter(Organization.is_deleted == False, Organization.is_platform_org == False).order_by(Organization.id.desc()).limit(15).all()
        for o in recent_orgs:
            joined = o.license_start_date.strftime('%Y-%m-%d') if o.license_start_date else '-'
            data_list.append([
                o.name or '-',
                o.industry or '-',
                o.subscription_plan or '-',
                o.subscription_status or '-',
                joined
            ])
            
    elif report_type == 'revenue':
        total_revenue = db.session.query(db.func.sum(SubscriptionPayment.final_amount)).filter_by(payment_status='Completed').scalar() or 0.0
        avg_rev = db.session.query(db.func.avg(SubscriptionPayment.final_amount)).filter_by(payment_status='Completed').scalar() or 0.0
        
        summary_metrics = [
            ["Total Completed Revenue", f"INR {total_revenue:,.2f}", "Financial Ledger Summary"],
            ["Average Payment Amount", f"INR {avg_rev:,.2f}", "Financial Transaction Average"]
        ]
        
        headers_list = ["Txn ID", "Organization", "Amount", "Cycle", "Status", "Date"]
        payments = SubscriptionPayment.query.filter_by(payment_status='Completed').order_by(SubscriptionPayment.created_at.desc()).limit(30).all()
        for p in payments:
            org_name = p.organization.name if p.organization else '-'
            date_str = p.created_at.strftime('%Y-%m-%d') if p.created_at else '-'
            data_list.append([
                f"TXN-{p.id}",
                org_name,
                f"INR {p.final_amount:,.2f}",
                p.billing_cycle or '-',
                p.payment_status or '-',
                date_str
            ])
            
    elif report_type in ('organizations', 'tenants'):
        headers_list = ["Organization Name", "Industry", "Admin Name", "Email", "Plan", "Status"]
        orgs = Organization.query.filter_by(is_deleted=False).order_by(Organization.created_at.desc() if hasattr(Organization, 'created_at') else Organization.id.desc()).limit(30).all()
        for o in orgs:
            data_list.append([
                o.name or '-',
                o.industry or '-',
                o.admin_name or '-',
                o.email or '-',
                o.subscription_plan or '-',
                o.subscription_status or '-'
            ])
            
    elif report_type == 'users':
        headers_list = ["Full Name", "Email", "Role", "Organization", "Status", "Created At"]
        users = User.query.order_by(User.created_at.desc()).limit(30).all()
        for u in users:
            org_name = u.organization.name if hasattr(u, 'organization') and u.organization else '-'
            role_name = u.role.name if hasattr(u, 'role') and u.role else '-'
            created = u.created_at.strftime('%Y-%m-%d') if u.created_at else '-'
            data_list.append([
                u.full_name or u.username or '-',
                u.email or '-',
                role_name,
                org_name,
                u.status or '-',
                created
            ])
            
    elif report_type == 'licenses':
        headers_list = ["License Number", "Organization", "Plan", "Max Users", "Storage (MB)", "Expires At"]
        orgs = Organization.query.filter_by(is_deleted=False).order_by(Organization.id.desc()).limit(30).all()
        for o in orgs:
            expires = o.license_expiry_date.strftime('%Y-%m-%d') if o.license_expiry_date else 'Lifetime'
            data_list.append([
                o.license_number or '-',
                o.name or '-',
                o.subscription_plan or '-',
                str(o.max_users or 0),
                f"{o.storage_limit_mb:,.0f} MB",
                expires
            ])
            
    elif report_type == 'subscriptions':
        headers_list = ["Sub ID", "Organization", "Plan", "Billing Cycle", "Amount", "Status", "Expires At"]
        subs = Subscription.query.order_by(Subscription.created_at.desc()).limit(30).all()
        for s in subs:
            org_name = s.organization.name if hasattr(s, 'organization') and s.organization else '-'
            expires = s.license_expiry_date.strftime('%Y-%m-%d') if hasattr(s, 'license_expiry_date') and s.license_expiry_date else 'Lifetime'
            data_list.append([
                f"SUB-{s.id}",
                org_name,
                s.subscription_plan or '-',
                s.billing_cycle or '-',
                f"INR {s.final_amount:,.2f}" if s.final_amount else '-',
                s.subscription_status or '-',
                expires
            ])
            
    elif report_type == 'modules':
        headers_list = ["Module Name", "Pricing Tier", "Status", "Creation Date"]
        modules = Module.query.limit(30).all()
        for m in modules:
            created = m.created_at.strftime('%Y-%m-%d') if hasattr(m, 'created_at') and m.created_at else '-'
            data_list.append([
                m.name or '-',
                m.pricing_tier or '-',
                m.status or '-',
                created
            ])
            
    elif report_type == 'support':
        headers_list = ["Ticket ID", "Organization", "Subject", "Priority", "Status", "Created At"]
        tickets = SupportTicket.query.order_by(SupportTicket.created_at.desc()).limit(30).all()
        for t in tickets:
            org_name = t.organization.name if hasattr(t, 'organization') and t.organization else '-'
            created = t.created_at.strftime('%Y-%m-%d') if t.created_at else '-'
            data_list.append([
                f"TCK-{t.id}",
                org_name,
                t.subject or '-',
                t.priority or '-',
                t.status or '-',
                created
            ])
            
    else:
        # Default empty fallback
        headers_list = ["Parameter", "Information Value"]
        data_list.append(["Scope", f"{report_type.capitalize()} Scope Ledger"])
        data_list.append(["Status", "Platform Operational"])

    b_ctx = DocumentBrandingService.get_branding_context(user.org_id if user else None)
    b_tmpl = DocumentBrandingService.get_template_config('analytics', user.org_id if user else None)

    # ─── PDF Format Generator ───
    if fmt in ('PDF', 'PRINT'):
        from fpdf import FPDF
        
        class QCMS_Analytics_PDF(FPDF):
            def header(self):
                self.set_fill_color(15, 23, 42)
                self.rect(0, 0, 210, 24, 'F')
                
                self.set_font('Helvetica', 'B', 12)
                self.set_text_color(255, 255, 255)
                self.set_xy(10, 7)
                self.cell(0, 10, clean_str(b_ctx['software_display_name'].upper()), 0, 0, 'L')
                
                self.set_font('Helvetica', 'I', 8)
                self.set_xy(10, 7)
                self.cell(190, 10, clean_str(b_tmpl['subtitle'].upper()), 0, 0, 'R')
                
                self.set_text_color(0, 0, 0)
                self.set_draw_color(226, 232, 240)
                self.set_xy(10, 30)

            def footer(self):
                self.set_y(-15)
                self.set_font('Helvetica', 'I', 8)
                self.set_text_color(100, 116, 139)
                self.cell(0, 10, clean_str(f'Page {self.page_no()} of {{nb}}'), 0, 0, 'C')
                self.set_x(10)
                self.cell(0, 10, clean_str(b_tmpl['footer_text']), 0, 0, 'L')
                self.set_x(-60)
                self.cell(50, 10, clean_str(datetime.now(timezone.utc).replace(tzinfo=None).strftime('%Y-%m-%d %H:%M:%S UTC')), 0, 0, 'R')

        pdf = QCMS_Analytics_PDF()
        pdf.alias_nb_pages()
        pdf.add_page()
        
        # Report Title
        pdf.set_font('Helvetica', 'B', 16)
        pdf.set_text_color(30, 41, 59)
        pdf.cell(0, 10, clean_str(f"{b_ctx['software_name']}: {report_type.upper()} {b_tmpl['header_title'].upper()}"), 0, 1, 'L')
        pdf.ln(2)
        
        # Metadata Block
        pdf.set_font('Helvetica', '', 9)
        pdf.set_text_color(100, 116, 139)
        pdf.cell(0, 5, clean_str(f"Document ID: {b_ctx['software_short_name']}-AR-{timestamp}-{report_type.upper()}"), 0, 1)
        pdf.cell(0, 5, clean_str(f"Generated At: {datetime.now(timezone.utc).replace(tzinfo=None).strftime('%B %d, %Y %H:%M:%S UTC')}"), 0, 1)
        pdf.cell(0, 5, clean_str(f"Organization: {b_ctx['legal_company_name']} | Classification: {b_tmpl['confidential_text']}"), 0, 1)
        pdf.line(10, pdf.get_y() + 4, 200, pdf.get_y() + 4)
        pdf.ln(8)
        
        # 1. Draw Summary Metrics Table if available
        if summary_metrics:
            pdf.set_font('Helvetica', 'B', 11)
            pdf.set_text_color(30, 41, 59)
            pdf.cell(0, 8, clean_str("Executive Key Performance Indicators (KPIs):"), 0, 1)
            pdf.ln(2)
            
            pdf.set_font('Helvetica', 'B', 9)
            pdf.set_fill_color(30, 41, 59)
            pdf.set_text_color(255, 255, 255)
            
            widths = [80, 50, 60]
            pdf.cell(widths[0], 8, clean_str("Metric Description"), 1, 0, 'L', True)
            pdf.cell(widths[1], 8, clean_str("Live Value"), 1, 0, 'R', True)
            pdf.cell(widths[2], 8, clean_str("Scope Segment"), 1, 1, 'C', True)
            
            pdf.set_font('Helvetica', '', 9)
            pdf.set_text_color(51, 65, 85)
            pdf.set_fill_color(248, 250, 252)
            fill = False
            for m in summary_metrics:
                pdf.cell(widths[0], 8, clean_str(m[0]), 'LRB', 0, 'L', fill)
                pdf.cell(widths[1], 8, clean_str(m[1]), 'RB', 0, 'R', fill)
                pdf.cell(widths[2], 8, clean_str(m[2]), 'RB', 1, 'C', fill)
                fill = not fill
            pdf.ln(8)
            
        # 2. Draw Main Tab Data Table
        pdf.set_font('Helvetica', 'B', 11)
        pdf.set_text_color(30, 41, 59)
        sec_title = "Roster Data Directory Ledger:" if report_type != 'overview' else "Platform Registered Subdivisions:"
        pdf.cell(0, 8, clean_str(sec_title), 0, 1)
        pdf.ln(2)
        
        if not data_list:
            pdf.set_font('Helvetica', 'I', 9)
            pdf.set_text_color(100, 116, 139)
            pdf.cell(0, 10, clean_str("No record entries located inside this platform catalog currently."), 0, 1)
        else:
            # Calculate column widths to distribute exactly across 190mm printable width
            num_cols = len(headers_list)
            # Custom column widths for better alignment
            col_widths = []
            if report_type == 'overview':
                col_widths = [60, 40, 30, 30, 30] # total 190
            elif report_type == 'revenue':
                col_widths = [25, 60, 30, 25, 25, 25] # total 190
            elif report_type in ('organizations', 'tenants'):
                col_widths = [50, 30, 35, 45, 15, 15] # total 190
            elif report_type == 'users':
                col_widths = [40, 45, 25, 45, 15, 20] # total 190
            elif report_type == 'licenses':
                col_widths = [40, 50, 30, 20, 25, 25] # total 190
            elif report_type == 'subscriptions':
                col_widths = [25, 50, 30, 25, 30, 15, 15] # total 190
            elif report_type == 'modules':
                col_widths = [60, 45, 40, 45] # total 190
            elif report_type == 'support':
                col_widths = [25, 50, 60, 20, 15, 20] # total 190
            else:
                col_widths = [95, 95]
                
            # Table Headers
            pdf.set_font('Helvetica', 'B', 8.5)
            pdf.set_fill_color(30, 41, 59)
            pdf.set_text_color(255, 255, 255)
            for i, h in enumerate(headers_list):
                pdf.cell(col_widths[i], 8, clean_str(h), 1, 0, 'C', True)
            pdf.ln(8)
            
            # Table Body
            pdf.set_font('Helvetica', '', 8)
            pdf.set_text_color(51, 65, 85)
            pdf.set_fill_color(248, 250, 252)
            fill = False
            for row in data_list:
                if pdf.get_y() > 265: # page break safeguard
                    pdf.add_page()
                    # Re-render headers
                    pdf.set_font('Helvetica', 'B', 8.5)
                    pdf.set_fill_color(30, 41, 59)
                    pdf.set_text_color(255, 255, 255)
                    for i, h in enumerate(headers_list):
                        pdf.cell(col_widths[i], 8, clean_str(h), 1, 0, 'C', True)
                    pdf.ln(8)
                    pdf.set_font('Helvetica', '', 8)
                    pdf.set_text_color(51, 65, 85)
                    pdf.set_fill_color(248, 250, 252)
                
                # Render cells
                for i, val in enumerate(row):
                    align = 'C'
                    if i == 0 or (report_type in ('overview', 'organizations', 'tenants', 'users', 'licenses', 'subscriptions', 'support') and i == 1):
                        align = 'L'
                    if 'INR' in str(val) or 'MB' in str(val) or (report_type == 'revenue' and i == 2):
                        align = 'R'
                    
                    border_style = 'LRB' if i == 0 else ('RB' if i == len(row)-1 else 'RB')
                    val_str = str(val)
                    if len(val_str) > 40:
                        val_str = val_str[:37] + "..."
                        
                    pdf.cell(col_widths[i], 7.5, clean_str(val_str), border_style, 0, align, fill)
                pdf.ln(7.5)
                fill = not fill
                
        # 3. Compliance and Security Section
        pdf.ln(8)
        if pdf.get_y() > 240:
            pdf.add_page()
        pdf.set_font('Helvetica', 'B', 10)
        pdf.set_text_color(30, 41, 59)
        pdf.cell(0, 8, clean_str("System Audit Verification & Security Disclaimer:"), 0, 1)
        pdf.set_font('Helvetica', '', 8)
        pdf.set_text_color(71, 85, 105)
        pdf.multi_cell(0, 4, clean_str("All analytical metrics, database rows, user activities, and subscription records printed above are generated directly from the live production database in real time. Access to this document is audited, tracked, and cryptographically verified under the central system governance policy. Disseminating confidential database statistics to external environments is strictly prohibited."))
        
        pdf_out = pdf.output(dest='S')
        if isinstance(pdf_out, str):
            pdf_bytes = pdf_out.encode('latin-1')
        else:
            pdf_bytes = pdf_out
            
        return send_file(
            io.BytesIO(pdf_bytes),
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f"export_{report_type}_{timestamp}.pdf"
        )
        
    # ─── Excel Format Generator (real XLSX via openpyxl) ───
    elif fmt == 'EXCEL':
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            wb = Workbook()
            ws = wb.active
            ws.title = report_type.capitalize()

            # Styles
            hdr_font   = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
            hdr_fill   = PatternFill('solid', fgColor='0F172A')
            hdr_align  = Alignment(horizontal='center', vertical='center', wrap_text=True)
            meta_font  = Font(name='Calibri', bold=True, size=10, color='1E293B')
            thin_side  = Side(style='thin', color='CBD5E1')
            thin_border= Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
            alt_fill   = PatternFill('solid', fgColor='F8FAFC')
            body_font  = Font(name='Calibri', size=10)
            body_align = Alignment(vertical='center')

            # ── Metadata rows ──
            ws.append([f"{b_ctx['software_name']} — {b_tmpl['header_title']} ({report_type.upper()})"])
            ws.append(["Report Segment", report_type.upper()])
            ws.append(["Generated At (UTC)", datetime.now(timezone.utc).replace(tzinfo=None).strftime('%Y-%m-%d %H:%M:%S')])
            ws.append([])

            for r_idx in range(1, 4):
                ws.cell(r_idx, 1).font = meta_font

            # ── KPI Summary table ──
            if summary_metrics:
                kpi_start = ws.max_row + 1
                ws.append(["Executive KPI Summary", "", ""])
                ws.cell(ws.max_row, 1).font = Font(name='Calibri', bold=True, size=11, color='1E293B')

                ws.append(["Metric Description", "Live Value", "Data Classification"])
                kpi_hdr_row = ws.max_row
                for col in range(1, 4):
                    cell = ws.cell(kpi_hdr_row, col)
                    cell.font    = hdr_font
                    cell.fill    = hdr_fill
                    cell.alignment = hdr_align
                    cell.border  = thin_border

                for idx, m in enumerate(summary_metrics):
                    ws.append(m[:3] if len(m) >= 3 else m + [''] * (3 - len(m)))
                    row_no = ws.max_row
                    fill = alt_fill if idx % 2 == 0 else None
                    for col in range(1, 4):
                        cell = ws.cell(row_no, col)
                        cell.font      = body_font
                        cell.alignment = body_align
                        cell.border    = thin_border
                        if fill:
                            cell.fill = fill
                ws.append([])

            # ── Main data table ──
            if headers_list:
                ws.append(["Detailed Catalog Entries"])
                ws.cell(ws.max_row, 1).font = Font(name='Calibri', bold=True, size=11, color='1E293B')

                ws.append(headers_list)
                data_hdr_row = ws.max_row
                for col in range(1, len(headers_list) + 1):
                    cell = ws.cell(data_hdr_row, col)
                    cell.font      = hdr_font
                    cell.fill      = hdr_fill
                    cell.alignment = hdr_align
                    cell.border    = thin_border

                for idx, row in enumerate(data_list):
                    ws.append([str(v) for v in row])
                    row_no = ws.max_row
                    fill = alt_fill if idx % 2 == 0 else None
                    for col in range(1, len(row) + 1):
                        cell = ws.cell(row_no, col)
                        cell.font      = body_font
                        cell.alignment = body_align
                        cell.border    = thin_border
                        if fill:
                            cell.fill = fill

            # ── Auto-size columns ──
            for col_cells in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col_cells[0].column)
                for cell in col_cells:
                    try:
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                    except Exception:
                        pass
                ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 50)

            xlsx_buf = io.BytesIO()
            wb.save(xlsx_buf)
            xlsx_buf.seek(0)

            return send_file(
                xlsx_buf,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=f"export_{report_type}_{timestamp}.xlsx"
            )
        except Exception as exc:
            import traceback; traceback.print_exc()
            return internal_server_error(exc, "Excel generation failed.")
        
    # ─── CSV Format Generator ───
    elif fmt == 'CSV':
        import csv as _csv
        si = io.StringIO()
        cw = _csv.writer(si)
        cw.writerow([f"{b_ctx['software_name']} - {b_tmpl['header_title']} ({report_type.upper()})"])
        cw.writerow(["Report Segment", report_type.upper()])
        cw.writerow(["Generated At (UTC)", datetime.now(timezone.utc).replace(tzinfo=None).strftime('%Y-%m-%d %H:%M:%S')])
        cw.writerow([])

        if summary_metrics:
            cw.writerow(["Executive KPI Summary"])
            cw.writerow(["Metric Description", "Live Value", "Data Classification"])
            for m in summary_metrics:
                cw.writerow(m)
            cw.writerow([])

        if headers_list and data_list:
            cw.writerow(["Detailed Catalog Entries"])
            cw.writerow(headers_list)
            for row in data_list:
                cw.writerow([str(v) for v in row])

        output = si.getvalue()
        from flask import Response as _Response
        return _Response(
            output,
            mimetype="text/csv",
            headers={
                "Content-Disposition": f"attachment;filename=export_{report_type}_{timestamp}.csv",
                "Access-Control-Expose-Headers": "Content-Disposition"
            }
        )

    else:
        return jsonify({"error": "Unsupported export format"}), 400
